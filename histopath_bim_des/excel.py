"""Functions for reading Excel data"""

from datetime import datetime
from os import PathLike
from warnings import warn

import openpyxl as oxl
import pandas as pd
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font
from openpyxl.utils import range_boundaries
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.filters import AutoFilter, FilterColumn
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

CellType = int | float | str | datetime


def get_name(wbook: oxl.Workbook, name: str) -> CellType | list[list[CellType]]:
    """Read an Excel named range as a single value or NumPy array.
    Arrays are flattened if possible.

    Args:
        wbook: The workbook object.
        name: Name of the Excel range to read.

    Returns: A single value, or a nested list containing the named range's values.
    """

    worksheet, cell_range = list(wbook.defined_names[name].destinations)[0]
    cell_range = str.replace(cell_range, '$', '')  # strip $ from cell range string
    cells = wbook[worksheet][cell_range]
    if isinstance(cells, Cell):
        return cells.value
    return [[cell.value for cell in r] for r in cells]


def get_table(wbook: oxl.Workbook, sheet_name: str, name: str) -> list[list[CellType]]:
    """Reads an Excel table as a nested list.

    Args:
        wbook: The workbook object.
        sheet_name: Name of the Excel worksheet containing the table.
        name: Name of the Excel table to read.

    Returns: A nested list containing the named range's values.
    """
    worksheet = wbook[sheet_name]
    cell_range = worksheet[worksheet.tables[name].ref]
    return [[cell.value for cell in r] for r in cell_range]


def remove_table_filters(table: Table, ws: Worksheet) -> None:
    """Remove filters from an Excel table.

    Args:
        table: The table to alter.
        ws: The worksheet in which the table is to be inserted.
    """
    # https://stackoverflow.com/questions/62212940/

    # Only remove filters if there is a header
    if not table.headerRowCount:
        return

    # Initialize table if not done (from WorksheetWriter.write_tables)
    if not table.tableColumns:
        min_col, _, max_col, _ = range_boundaries(table.ref)
        for idx in range(min_col, max_col + 1):
            col = TableColumn(id=idx, name=f"Column{idx}")
            table.tableColumns.append(col)
        table.autoFilter = AutoFilter(ref=table.ref)
        try:
            row = ws[table.ref][0]
            for cell, col in zip(row, table.tableColumns):
                if cell.data_type != "s":
                    warn("File may not be readable: column headings must be strings.")
                col.name = str(cell.value)
        except TypeError:
            warn("Column headings are missing, file may not be readable")

    filter_columns = table.autoFilter.filterColumn
    # clear any current FilterColumns that might be there
    filter_columns.clear()
    for idx in range(len(table.tableColumns)):
        # ColId of filters is relative to Table, so starts at 0
        # Remove filter with hiddenButton=True
        filter_columns.append(FilterColumn(colId=idx, hiddenButton=True, filters=None))


def write_table(
    df: pd.DataFrame,
    path: PathLike,
    sheet_name: str,
    name: str
) -> None:
    """Writes a pandas DataFrame to an Excel table in a new worksheet, **replacing**
    the existing worksheet if found.  Note that the index is not used;
    use `df.reset_index(drop=False, names='idx')` to include the DataFrame index.

    Args:
        df: The DataFrame to write.
        path: The filepath of the workbook to write to.
        sheet_name: The name of the worksheet to write to.
        name: The name of the Excel table to create.
    """

    # Remove worksheet if exists.
    # For some reason, we need to save and re-open the worksheet to properly
    # remove the existing table.
    wbook = oxl.load_workbook(path)
    if sheet_name in wbook.sheetnames:
        # print(f'Deleting existing worksheet {sheet_name}')
        del wbook[sheet_name]
        wbook.save(path)
    wbook.close()

    wbook2 = oxl.load_workbook(path)
    wbook2.create_sheet(sheet_name)
    ws = wbook2[sheet_name]

    for r, v in enumerate(dataframe_to_rows(df, index=False, header=True), start=2):
        for c, w in enumerate(v, start=1):
            ws.cell(r, c, value=w)

    ft = Font(
        bold=True,
        size=18
    )
    ws.cell(1, 1, value=sheet_name)
    ws['A1'].font = ft

    tab = Table(displayName=name, ref=f"$A$2:$B${len(df.index)+2}")
    style = TableStyleInfo(
        name='TableStyleLight9',
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    remove_table_filters(tab, ws)

    ws.add_table(tab)
    wbook2.save(path)
    wbook2.close()
